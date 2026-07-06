"""Exportación de tablas del dashboard a Excel con formato profesional."""
import io
import pandas as pd
import streamlit as st

from domain.formato import texto_a_numero


def df_to_excel_bytes(df: pd.DataFrame, money_cols=None) -> bytes:
    """Exporta a Excel con estilo profesional. Las columnas en `money_cols`
    (que en pantalla llegan como texto '$1,234.56' para que se vean bien en
    st.dataframe) se convierten de vuelta a número real con formato de
    moneda de Excel — así el archivo exportado se puede sumar, ordenar y
    graficar directamente en Excel, en vez de quedar como texto."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    money_cols = set(money_cols or [])
    df2 = df.copy()
    for c in money_cols:
        if c in df2.columns:
            df2[c] = df2[c].apply(texto_a_numero)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df2.to_excel(writer, index=False, sheet_name="Datos")
        ws = writer.sheets["Datos"]

        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        border_fina = Border(bottom=Side(style="thin", color="D9D9D9"))

        for col_idx, col_name in enumerate(df2.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            es_moneda = col_name in money_cols
            max_len = len(str(col_name))
            for row_idx in range(2, ws.max_row + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.border = border_fina
                if es_moneda:
                    c.number_format = '$#,##0.00;[Red]-$#,##0.00'
                    c.alignment = Alignment(horizontal="right")
                val_len = (len(f"{c.value:,.2f}") if (es_moneda and isinstance(c.value, (int, float)))
                           else len(str(c.value)) if c.value is not None else 0)
                max_len = max(max_len, val_len)
            ws.column_dimensions[cell.column_letter].width = min(max_len + 3, 42)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    return buf.getvalue()


def download_button_excel(df: pd.DataFrame, filename: str, label: str = "⬇ Descargar Excel", money_cols=None):
    data = df_to_excel_bytes(df, money_cols=money_cols)
    st.download_button(label=label, data=data, file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
